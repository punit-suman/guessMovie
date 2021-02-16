import openpyxl
import random


def getMovieDetails():
    wb = openpyxl.load_workbook('movies.xlsx')
    sheet = wb['Sheet1']
    index = random.randint(2, sheet.max_row)
    cell = sheet.cell(index, 1)
    movie_name = cell.value
    cell = sheet.cell(index, 2)
    hint1 = cell.value
    cell = sheet.cell(index, 3)
    hint21 = cell.value
    cell = sheet.cell(index, 4)
    hint22 = cell.value
    return (movie_name, hint1, hint21, hint22)